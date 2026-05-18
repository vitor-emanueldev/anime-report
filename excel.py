import openpyxl
import logging
from openpyxl.styles import PatternFill, Font, Alignment

logger = logging.getLogger(__name__)

def gerar_excel(animes):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["Título", "Nota", "Episódios", "Status", "Gêneros", "Sinopse"])

    for anime in animes:
        genres = ", ".join([g["name"] for g in anime["genres"]])
        title = anime["title_english"] or anime["title"]
        score = anime["score"] or "N/A"
        episodes = anime["episodes"] or "N/A"
        status = anime["status"]
        synopsis = anime["synopsis"]
        ws.append([title, score, episodes, status, genres, synopsis])
    cabecalho_fill = PatternFill(start_color="2F75B6", end_color="2F75B6", fill_type="solid")
    cabecalho_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = cabecalho_fill
        cell.font = cabecalho_font
        cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

    fill_claro = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if i % 2 == 0:
            for cell in row:
                cell.fill = fill_claro

    wb.save("animes.xlsx")
    logger.info("Arquivo animes.xlsx gerado com sucesso!")

